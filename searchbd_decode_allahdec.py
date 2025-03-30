#Декоднул @allahdec.
__import__("webbrowser").open("https://t.me/+-aF7ZoooDo5mYmVi")


import os  #Декоднул @allahdec.
import platform  #Декоднул @allahdec.
import csv  #Декоднул @allahdec.
import socket  #Декоднул @allahdec.
import fake_useragent  #Декоднул @allahdec.
import requests  #Декоднул @allahdec.
import termcolor  #Декоднул @allahdec.
import pyfiglet  #Декоднул @allahdec.
from fake_useragent import UserAgent  #Декоднул @allahdec.
from termcolor import colored  #Декоднул @allahdec.

os.system('cls' if os.name == 'nt' else 'clear')

banner = '\n ███▄ ▄███▓▓██   ██▓  ██████ ▄▄▄█████▓▓█████  ██▀███ ▓██   ██▓\n▓██▒▀█▀ ██▒ ▒██  ██▒▒██    ▒ ▓  ██▒ ▓▒▓█   ▀ ▓██ ▒ ██▒▒██  ██▒\n▓██    ▓██░  ▒██ ██░░ ▓██▄   ▒ ▓██░ ▒░▒███   ▓██ ░▄█ ▒ ▒██ ██░\n▒██    ▒██   ░ ▐██▓░  ▒   ██▒░ ▓██▓ ░ ▒▓█  ▄ ▒██▀▀█▄   ░ ▐██▓░\n▒██▒   ░██▒  ░ ██▒▓░▒██████▒▒  ▒██▒ ░ ░▒████▒░██▓ ▒██▒ ░ ██▒▓░\n░ ▒░   ░  ░   ██▒▒▒ ▒ ▒▓▒ ▒ ░  ▒ ░░   ░░ ▒░ ░░ ▒▓ ░▒▓░  ██▒▒▒ \n░  ░      ░ ▓██ ░▒░ ░ ░▒  ░ ░    ░     ░ ░  ░  ░▒ ░ ▒░▓██ ░▒░ \n░      ░    ▒ ▒ ░░  ░  ░  ░    ░         ░     ░░   ░ ▒ ▒ ░░  \n       ░    ░ ░           ░              ░  ░   ░     ░ ░     \n            ░ ░                                       ░ ░     \n'

print(colored(banner, 'magenta'))

def get_directory_size(directory):  #Декоднул @allahdec.
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(directory):  #Декоднул @allahdec.
        for f in filenames:  #Декоднул @allahdec.
            fp = os.path.join(dirpath, f)
            total_size += os.path.getsize(fp)
    return total_size  #Декоднул @allahdec.

def format_size(size):  #Декоднул @allahdec.
    if size < 1048576:  #Декоднул @allahdec.
        return f"{size} байт"  #Декоднул @allahdec.
    elif size < 1073741824:  #Декоднул @allahdec.
        return f"{size / 1048576:.2f} МБ"  #Декоднул @allahdec.
    else:  #Декоднул @allahdec.
        return f"{size / 1073741824:.2f} ГБ"  #Декоднул @allahdec.

def search_in_file(file_path, keyword):  #Декоднул @allahdec.
    try:  #Декоднул @allahdec.
        if file_path.endswith('.csv'):  #Декоднул @allahdec.
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:  #Декоднул @allahdec.
                reader = csv.reader(f)
                for row in reader:  #Декоднул @allahdec.
                    if any(keyword in str(cell) for cell in row):  #Декоднул @allahdec.
                        print(colored(f"Found in {file_path}: {row}", 'magenta'))
        elif file_path.endswith(('.xlsx', '.xls')):  #Декоднул @allahdec.
            import openpyxl  #Декоднул @allahdec.
            load_workbook = openpyxl.load_workbook
            wb = load_workbook(file_path, data_only=True)
            for sheet in wb.sheetnames:  #Декоднул @allahdec.
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):  #Декоднул @allahdec.
                    if any(keyword in str(cell) for cell in row):  #Декоднул @allahdec.
                        print(colored(f"Found in {file_path} ({sheet}): {row}", 'magenta'))
        elif file_path.endswith(('.txt', '.sql')):  #Декоднул @allahdec.
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:  #Декоднул @allahdec.
                for line in f:  #Декоднул @allahdec.
                    if keyword in line:  #Декоднул @allahdec.
                        print(colored(f"Found in {file_path}: {line.strip()}", 'magenta'))
        else:  #Декоднул @allahdec.
            return None  #Декоднул @allahdec.
    except Exception as e:  #Декоднул @allahdec.
        print(colored(f"Error processing {file_path}: {e}", 'magenta'))
        return None  #Декоднул @allahdec.

def search_keyword_in_directory(directory, keyword):  #Декоднул @allahdec.
    for filename in os.listdir(directory):  #Декоднул @allahdec.
        file_path = os.path.join(directory, filename)
        if os.path.isfile(file_path):  #Декоднул @allahdec.
            search_in_file(file_path, keyword)

def main():  #Декоднул @allahdec.
    db_directory = 'bases'
    db_size = get_directory_size(db_directory)
    formatted_size = format_size(db_size)
    print(colored(f'Размер баз данных: {formatted_size}', 'magenta'))

    mennu = '╔══════════════════════════╗   \n║[1] - Начать поиск        ║\n║[99] - Выход              ║\n╚══════════════════════════╝'
    while True:  #Декоднул @allahdec.
        print(colored(mennu, 'magenta'))
        choice = input('Выберите пункт: ').strip()

        if choice == '1':  #Декоднул @allahdec.
            print('Введите запрос: ')
            keyword = input().strip()
            print(colored(f"Поиск в папке '{db_directory}'...", 'magenta'))
            search_keyword_in_directory(db_directory, keyword)
            input('Нажмите Enter, чтобы продолжить...')
        elif choice == '99':  #Декоднул @allahdec.
            print('Выход из программы.')
            os.system('python main.py')
            return None  #Декоднул @allahdec.
        else:  #Декоднул @allahdec.
            print('Неверный выбор.')
            input('Нажмите Enter, чтобы продолжить...')

if __name__ == '__main__':  #Декоднул @allahdec.
    main()
